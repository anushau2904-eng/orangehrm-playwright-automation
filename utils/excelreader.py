from openpyxl import load_workbook
from pathlib import Path   


class ExcelReader:

    PROJECT_ROOT = Path(__file__).resolve().parent.parent

    @classmethod
    def get_user_data(cls, file_name):

        file_path = (
            cls.PROJECT_ROOT
            / "testdata"
            / file_name
        )

        workbook = load_workbook(file_path)

        sheet = workbook.active

        headers = [cell.value for cell in sheet[1]]

        users = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            users.append(dict(zip(headers, row)))

        workbook.close()

        return users